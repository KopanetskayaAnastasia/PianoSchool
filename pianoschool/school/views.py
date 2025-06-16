from django.views.generic import ListView

from .forms import ProgressReportExportForm
from .models import Lesson, Subject, EducationPlan, Test
from users.models import Students, Teachers, TestResult, TeacherStudent
from django.db.models import F
from django.db.models.functions import Cast
from django.db.models import IntegerField
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import View, DetailView
from school.models import School
from django.contrib import messages


class SolfeggioLessonsView(LoginRequiredMixin, ListView):
    model = Lesson
    template_name = 'school/solfeggio_lessons.html'
    context_object_name = 'lessons'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return render(request, "404.html", status=404)
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        user = self.request.user
        # Если пользователь учитель — показываем уроки по сольфеджио для его школы
        if user.groups.filter(name="Учителя").exists():
            try:
                teacher = Teachers.objects.get(user=user)
            except Teachers.DoesNotExist:
                return Lesson.objects.none()
            school = teacher.school
            # Находим предмет "Сольфеджио"
            try:
                solfeggio_subject = Subject.objects.get(name='Сольфеджио')
            except Subject.DoesNotExist:
                return Lesson.objects.none()
            # Находим все планы обучения по сольфеджио для школы учителя
            plans = EducationPlan.objects.filter(school=school,subject=solfeggio_subject)
            lessons = Lesson.objects.filter(education_plan__in=plans)
            lessons = lessons.annotate(
                json_id=Cast(F('content_file__id'),output_field=IntegerField())
            ).order_by('json_id')
            return lessons

        # Если пользователь ученик — показываем только его уроки
        try:
            student = Students.objects.get(user=user)
        except Students.DoesNotExist:
            return Lesson.objects.none()
        school = student.school
        grade = student.grade_of_school
        try:
            solfeggio_subject = Subject.objects.get(name='Сольфеджио')
        except Subject.DoesNotExist:
            return Lesson.objects.none()
        plans = EducationPlan.objects.filter(
            year=grade,
            school=school,
            subject=solfeggio_subject
        )
        lessons = Lesson.objects.filter(education_plan__in=plans)
        lessons = lessons.annotate(
            json_id=Cast(F('content_file__id'),output_field=IntegerField())
        ).order_by('json_id')

        return lessons

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Добавляем информацию о предмете
        try:
            context['subject'] = Subject.objects.get(name='Сольфеджио')
        except Subject.DoesNotExist:
            pass
        user = self.request.user
        # Если пользователь - ученик, добавляем информацию о его классе
        if user.groups.filter(name="Ученики").exists():
            try:
                student = Students.objects.get(user=user)
                context['student_grade'] = student.grade_of_school
                context['student_school'] = student.school
            except Students.DoesNotExist:
                pass
        # Если пользователь - учитель, добавляем информацию о его школе
        elif user.groups.filter(name="Учителя").exists():
            try:
                teacher = Teachers.objects.get(user=user)
                context['teacher_school'] = teacher.school
            except Teachers.DoesNotExist:
                pass
        return context


class LessonSolfDetailView(LoginRequiredMixin, DetailView):
    model = Lesson
    template_name = 'school/lesson_detail_solf.html'
    context_object_name = 'lesson'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return render(request, "404.html", status=404)
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        lesson = self.get_object()
        # Добавляем связанные тесты, если они есть
        context['tests'] = Test.objects.filter(lesson=lesson)
        # Получаем следующий урок
        try:
            context['next_lesson'] = lesson.get_next_lesson()
        except:
            pass

        return context


class TestDetailView(LoginRequiredMixin, View):
    template_name = 'school/test_detail.html'
    result_template = 'school/test_result.html'
    max_attempts = 2

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return render(request, "404.html", status=404)
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, pk):
        test = get_object_or_404(Test, pk=pk)
        user = request.user
        if user.groups.filter(name="Учителя").exists():
            # Учитель только смотрит тест
            context = {
                'test': test,
                'is_teacher': True,
            }
            return render(request, self.template_name, context)

        if user.groups.filter(name="Ученики").exists():
            try:
                student = Students.objects.get(user=user)
            except Students.DoesNotExist:
                messages.error(request, "Вы не зарегистрированы как ученик.")
                return redirect('school:solfeggio_lessons')
            # Проверяем, есть ли уже результат
            result = TestResult.objects.filter(student=student, test=test).first()
            attempts = request.session.get(f'test_{test.pk}_attempts', 0)
            if result:
                # Тест уже пройден и отправлен
                return render(request, self.template_name, {
                    'test': test,
                    'test_completed': True,
                    'result': result,
                    'is_teacher': False,
                })
            # Если попыток больше 2, не даём проходить
            if attempts >= self.max_attempts:
                return render(request, self.template_name, {
                    'test': test,
                    'attempts': attempts,
                    'no_more_attempts': True,
                    'is_teacher': False,
                })
            # Показываем форму теста
            return render(request, self.template_name, {
                'test': test,
                'attempts': attempts,
                'is_teacher': False,
            })
        # Не студент и не учитель
        messages.error(request, "Нет доступа к тесту.")
        return redirect('school:solfeggio_lessons')

    def post(self, request, pk):
        test = get_object_or_404(Test, pk=pk)
        user = request.user
        if not user.groups.filter(name="Ученики").exists():
            messages.error(request, "Только студенты могут проходить тесты.")
            return redirect('school:test_detail', pk=pk)
        try:
            student = Students.objects.get(user=user)
        except Students.DoesNotExist:
            messages.error(request, "Вы не зарегистрированы как ученик.")
            return redirect('school:solfeggio_lessons')
        attempts = request.session.get(f'test_{test.pk}_attempts', 0)
        if attempts >= self.max_attempts:
            return redirect('school:test_detail', pk=pk)
        # Проверяем, не отправлен ли уже результат
        if TestResult.objects.filter(student=student, test=test).exists():
            return redirect('school:test_detail', pk=pk)
        # Проверяем структуру теста
        test_json = test.content.get('test', {})
        questions = test_json.get('questions', [])
        correct_count = 0
        incorrect_count = 0
        score = 0
        max_score = sum(q.get('points', 0) for q in questions)
        user_answers = {}
        for idx, q in enumerate(questions):
            qid = q.get('id')
            user_answer = request.POST.get(f'question_{qid}')
            user_answers[str(qid)] = user_answer
            if user_answer and user_answer.strip() == q.get('correct_answer'):
                correct_count += 1
                score += q.get('points', 0)
            else:
                incorrect_count += 1
        # Сохраняем попытку в сессии (не в БД!)
        attempts += 1
        request.session[f'test_{test.pk}_attempts'] = attempts
        # Формируем данные для окна результатов
        passing_score = test_json.get('passing_score', 0)
        feedback_message = ""
        for rng in test_json.get('feedback', {}).get('ranges', []):
            if rng['min'] <= score <= rng['max']:
                feedback_message = rng['message']
                break
        # Сохраняем только сериализуемые данные в сессию
        session_result = {
            'score': score,
            'max_score': max_score,
            'correct_count': correct_count,
            'incorrect_count': incorrect_count,
            'passing_score': passing_score,
            'feedback_message': feedback_message,
            'attempts': attempts,
            'can_retry': attempts < self.max_attempts,
            'user_answers': user_answers,
        }
        request.session[f'test_{test.pk}_last_result'] = session_result
        # Для шаблона добавляем объект test только в context
        context = dict(session_result)
        context['test'] = test
        context['is_teacher'] = False

        return render(request, self.result_template, context)


class TestResultSubmitView(LoginRequiredMixin, View):
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return render(request, "404.html", status=404)
        return super().dispatch(request, *args, **kwargs)
    def post(self, request, pk):
        test = get_object_or_404(Test, pk=pk)
        user = request.user
        if not user.groups.filter(name="Ученики").exists():
            messages.error(request, "Только студенты могут отправлять результаты.")
            return redirect('school:test_detail', pk=pk)
        try:
            student = Students.objects.get(user=user)
        except Students.DoesNotExist:
            messages.error(request, "Вы не зарегистрированы как ученик.")
            return redirect('school:solfeggio_lessons')

        # Проверяем, не отправлен ли уже результат
        if TestResult.objects.filter(student=student, test=test).exists():
            return redirect('school:test_detail', pk=pk)

        # Получаем результат из сессии
        session_result = request.session.get(f'test_{test.pk}_last_result')
        if not session_result:
            messages.error(request, "Нет данных для отправки.")
            return redirect('school:test_detail', pk=pk)

        # Сохраняем результат в БД
        TestResult.objects.create(
            student=student,
            test=test,
            score=session_result['score'],
            correct_count=session_result['correct_count'],
            incorrect_count=session_result['incorrect_count'],
            max_score=session_result['max_score'],
            attempt_number=session_result['attempts'],
            details={'answers': session_result['user_answers']}
        )
        # Очищаем сессию
        del request.session[f'test_{test.pk}_last_result']
        messages.success(request, "Результаты успешно отправлены!")
        return redirect('school:test_detail', pk=pk)


@login_required
def student_progress(request):
    if not request.user.is_authenticated:
        return render(request, "404.html", status=404)
    if not request.user.groups.filter(name="Ученики").exists():
        return render(request, "school/forbidden.html")
    try:
        student = Students.objects.get(user=request.user)
    except Students.DoesNotExist:
        return render(request, "school/forbidden.html")

    # Сольфеджио
    solfeggio_results = TestResult.objects.filter(
        student=student,
        test__lesson__education_plan__subject__name="Сольфеджио"
    ).select_related('test', 'test__lesson').order_by('test__lesson__section', 'test__topic', '-timestamp')

    grouped_solfeggio = defaultdict(list)
    for res in solfeggio_results:
        section = res.test.lesson.section
        grouped_solfeggio[section].append(res)
    grouped_solfeggio_list = []
    for section in sorted(grouped_solfeggio.keys()):
        results = sorted(grouped_solfeggio[section], key=lambda r: r.test.topic)
        avg_score = round(sum(r.score for r in results) / len(results), 2) if results else 0
        grouped_solfeggio_list.append((section, results, avg_score))

    # Музыкальная литература
    literature_results = TestResult.objects.filter(
        student=student,
        test__lesson__education_plan__subject__name="Музыкальная литература"
    ).select_related('test', 'test__lesson').order_by('test__lesson__section', 'test__topic', '-timestamp')

    grouped_literature = defaultdict(list)
    for res in literature_results:
        section = res.test.lesson.section
        grouped_literature[section].append(res)
    grouped_literature_list = []
    for section in sorted(grouped_literature.keys()):
        results = sorted(grouped_literature[section], key=lambda r: r.test.topic)
        avg_score = round(sum(r.score for r in results) / len(results), 2) if results else 0
        grouped_literature_list.append((section, results, avg_score))

    # --- Вынесено из цикла: Ear Training ---
    ear_attempts = EarTrainingAttempt.objects.filter(user=request.user)
    task_ids = ear_attempts.values_list('task_id', flat=True)
    ear_tests = EarTrainingTest.objects.filter(tasks__id__in=task_ids).distinct()

    grouped_ear = defaultdict(list)
    for test in ear_tests:
        section = test.section
        test_tasks = test.tasks.all()
        test_attempts = ear_attempts.filter(task__in=test_tasks)
        correct_count = test_attempts.filter(is_correct=True).count()
        incorrect_count = test_attempts.filter(is_correct=False).count()
        score = sum(
            t.points for t in test_tasks
            if test_attempts.filter(task=t, is_correct=True).exists()
        )
        max_score = sum(t.points for t in test_tasks)
        last_attempt = test_attempts.order_by('-timestamp').first()
        grouped_ear[section].append({
            "test": test,
            "score": score,
            "correct_count": correct_count,
            "incorrect_count": incorrect_count,
            "max_score": max_score,
            "last_attempt": last_attempt.timestamp if last_attempt else None,
        })

    sorted_ear_sections = sorted(grouped_ear.keys())
    grouped_ear_training_results = []
    for section in sorted_ear_sections:
        results = grouped_ear[section]
        avg_score = round(sum(r["score"] for r in results) / len(results), 2) if results else 0
        grouped_ear_training_results.append((section, results, avg_score))

    return render(
        request,
        "school/student_progress.html",
        {
            "grouped_ear_training_results": grouped_ear_training_results,
            "grouped_solfeggio_results": grouped_solfeggio_list,
            "grouped_literature_results": grouped_literature_list,
        }
    )

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views import View
from django.contrib import messages

from .models import Group, GroupMembership
from .forms import GroupForm, GroupMembershipForm

class TeacherStudentsView(LoginRequiredMixin, View):
    template_name = "school/teacher_students.html"

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return render(request, "404.html", status=404)
        return super().dispatch(request, *args, **kwargs)

    def get(self, request):
        user = request.user
        if not user.groups.filter(name="Учителя").exists():
            return render(request, "school/forbidden.html")
        teacher = get_object_or_404(Teachers, user=user)

        # Получаем id всех учеников учителя
        teacher_student_ids = TeacherStudent.objects.filter(teacher=teacher).values_list('student_id', flat=True)

        # Получаем все школы, в которых есть ученики учителя
        school_ids = Students.objects.filter(pk__in=teacher_student_ids).values_list('school_id', flat=True).distinct()
        school_choices = School.objects.filter(id__in=school_ids)

        # Получаем все года обучения из EducationPlan для этих школ
        grade_choices = EducationPlan.objects.filter(
            school_id__in=school_ids
        ).values_list('year', flat=True).distinct().order_by('year')

        # Получаем фильтры из GET-параметров
        school_filter = request.GET.get('school_filter')
        grade_filter = request.GET.get('grade_filter')

        # Базовый queryset всех учеников учителя
        students_qs = Students.objects.filter(pk__in=teacher_student_ids)

        # Применяем фильтры
        if school_filter:
            students_qs = students_qs.filter(school_id=school_filter)
        if grade_filter:
            students_qs = students_qs.filter(grade_of_school=grade_filter)

        all_teacher_students = students_qs.select_related('school').order_by('grade_of_school', 'surname', 'name')

        # Группы учителя
        groups = Group.objects.filter(teacher=teacher).order_by('-created_at')

        # Текущая выбранная группа (по GET параметру)
        current_group_id = request.GET.get('group')
        current_group = None
        current_group_members = []
        if current_group_id:
            current_group = get_object_or_404(Group, pk=current_group_id, teacher=teacher)
            current_group_members = Students.objects.filter(
                pk__in=current_group.memberships.values_list('student_id', flat=True)
            ).select_related('school').order_by('grade_of_school', 'surname', 'name')

        # Форма создания новой группы, передаем отфильтрованный список учеников
        group_form = GroupForm()
        members_form = GroupMembershipForm(
            teacher=teacher,
            students_queryset=all_teacher_students
        )

        context = {
            "all_teacher_students": all_teacher_students,
            "groups": groups,
            "current_group": current_group,
            "current_group_members": current_group_members,
            "group_form": group_form,
            "members_form": members_form,
            "school_choices": school_choices,
            "grade_choices": list(grade_choices),
            "school_filter": school_filter or '',
            "grade_filter": grade_filter or '',
        }
        return render(request, self.template_name, context)

    def post(self, request):
        user = request.user
        if not user.groups.filter(name="Учителя").exists():
            return render(request, "school/forbidden.html")
        teacher = get_object_or_404(Teachers, user=user)

        # Получаем id всех учеников учителя
        teacher_student_ids = TeacherStudent.objects.filter(teacher=teacher).values_list('student_id', flat=True)

        # Получаем фильтры из GET-параметров (чтобы сохранить фильтрацию после POST)
        school_filter = request.GET.get('school_filter')
        grade_filter = request.GET.get('grade_filter')

        students_qs = Students.objects.filter(pk__in=teacher_student_ids)
        if school_filter:
            students_qs = students_qs.filter(school_id=school_filter)
        if grade_filter:
            students_qs = students_qs.filter(grade_of_school=grade_filter)

        # Создание новой группы
        if 'create_group' in request.POST:
            group_form = GroupForm(request.POST)
            members_form = GroupMembershipForm(request.POST, teacher=teacher, students_queryset=students_qs)
            if group_form.is_valid() and members_form.is_valid():
                group = group_form.save(commit=False)
                group.teacher = teacher
                group.save()
                for student in members_form.cleaned_data['students']:
                    GroupMembership.objects.create(group=group, student=student)
                messages.success(request, "Группа успешно создана.")
                return redirect(f"{request.path}?group={group.pk}&school_filter={school_filter or ''}&grade_filter={grade_filter or ''}")
            else:
                groups = Group.objects.filter(teacher=teacher).order_by('-created_at')
                all_teacher_students = students_qs.select_related('school').order_by('grade_of_school', 'surname', 'name')
                context = {
                    "all_teacher_students": all_teacher_students,
                    "groups": groups,
                    "current_group": None,
                    "current_group_members": [],
                    "group_form": group_form,
                    "members_form": members_form,
                    "school_choices": School.objects.filter(id__in=students_qs.values_list('school_id', flat=True).distinct()),
                    "grade_choices": list(grade_choices),
                    "school_filter": school_filter or '',
                    "grade_filter": grade_filter or '',
                }
                return render(request, self.template_name, context)

        # Редактирование существующей группы
        if 'edit_group_id' in request.POST:
            group_id = request.POST.get('edit_group_id')
            group = get_object_or_404(Group, pk=group_id, teacher=teacher)
            group_form = GroupForm(request.POST, instance=group)
            members_form = GroupMembershipForm(request.POST, teacher=teacher, students_queryset=students_qs)
            if group_form.is_valid() and members_form.is_valid():
                group_form.save()
                group.memberships.all().delete()
                for student in members_form.cleaned_data['students']:
                    GroupMembership.objects.create(group=group, student=student)
                messages.success(request, "Группа успешно обновлена.")
                return redirect(f"{request.path}?group={group.pk}&school_filter={school_filter or ''}&grade_filter={grade_filter or ''}")
            else:
                groups = Group.objects.filter(teacher=teacher).order_by('-created_at')
                all_teacher_students = students_qs.select_related('school').order_by('grade_of_school', 'surname', 'name')
                current_group_members = Students.objects.filter(
                    pk__in=group.memberships.values_list('student_id', flat=True)
                )
                context = {
                    "all_teacher_students": all_teacher_students,
                    "groups": groups,
                    "current_group": group,
                    "current_group_members": current_group_members,
                    "group_form": group_form,
                    "members_form": members_form,
                    "school_choices": School.objects.filter(id__in=students_qs.values_list('school_id', flat=True).distinct()),
                    "grade_choices": list(grade_choices),
                    "school_filter": school_filter or '',
                    "grade_filter": grade_filter or '',
                }
                return render(request, self.template_name, context)

        return redirect('school:teacher_students')


def get_grouped_ear_training_results(student):
    # Получаем все EarTrainingTest, где есть попытки у ученика
    attempts = EarTrainingAttempt.objects.filter(user=student.user)
    tests = EarTrainingTest.objects.filter(tasks__attempts__user=student.user).distinct()
    grouped = defaultdict(list)
    for test in tests:
        section = test.section
        test_tasks = test.tasks.all()
        test_attempts = attempts.filter(task__in=test_tasks)
        correct_count = test_attempts.filter(is_correct=True).count()
        incorrect_count = test_attempts.filter(is_correct=False).count()
        score = sum(
            t.points for t in test_tasks
            if test_attempts.filter(task=t, is_correct=True).exists()
        )
        max_score = sum(t.points for t in test_tasks)
        last_attempt = test_attempts.aggregate(Max('timestamp'))['timestamp__max']
        grouped[section].append({
            "title": test.title,
            "score": score,
            "correct_count": correct_count,
            "incorrect_count": incorrect_count,
            "max_score": max_score,
            "last_attempt": last_attempt,
        })
    return grouped


from collections import defaultdict
class StudentProgressDetailView(LoginRequiredMixin, DetailView):
    model = Students
    template_name = "school/student_progress_detail.html"
    context_object_name = "student"

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return render(request, "404.html", status=404)
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student = self.object

        # Сольфеджио
        solfeggio_results = TestResult.objects.filter(
            student=student,
            test__lesson__education_plan__subject__name="Сольфеджио"
        ).select_related('test', 'test__lesson').order_by('test__lesson__section', 'test__topic')

        grouped_solfeggio = defaultdict(list)
        for res in solfeggio_results:
            section = res.test.lesson.section
            grouped_solfeggio[section].append(res)
        sorted_sections = sorted(grouped_solfeggio.keys())
        grouped_solfeggio_list = []
        for section in sorted_sections:
            results = grouped_solfeggio[section]
            avg_score = round(sum(r.score for r in results) / len(results), 2) if results else 0
            grouped_solfeggio_list.append((section, results, avg_score))

        # Музыкальная литература
        literature_results = TestResult.objects.filter(
            student=student,
            test__lesson__education_plan__subject__name="Музыкальная литература"
        ).select_related('test', 'test__lesson').order_by('test__lesson__section', 'test__topic')

        grouped_literature = defaultdict(list)
        for res in literature_results:
            section = res.test.lesson.section
            grouped_literature[section].append(res)
        sorted_lit_sections = sorted(grouped_literature.keys())
        grouped_literature_list = []
        for section in sorted_lit_sections:
            results = grouped_literature[section]
            avg_score = round(sum(r.score for r in results) / len(results), 2) if results else 0
            grouped_literature_list.append((section, results, avg_score))

        # Получаем все попытки ученика по слуховым диктантам
        ear_attempts = EarTrainingAttempt.objects.filter(user=student.user)
        # Получаем id всех задач, по которым были попытки
        task_ids = ear_attempts.values_list('task_id', flat=True)
        # Получаем все контрольные, в которых есть эти задачи
        ear_tests = EarTrainingTest.objects.filter(tasks__id__in=task_ids).distinct()

        grouped_ear = defaultdict(list)
        for test in ear_tests:
            section = test.section
            test_tasks = test.tasks.all()
            test_attempts = ear_attempts.filter(task__in=test_tasks)
            correct_count = test_attempts.filter(is_correct=True).count()
            incorrect_count = test_attempts.filter(is_correct=False).count()
            score = sum(
                t.points for t in test_tasks
                if test_attempts.filter(task=t, is_correct=True).exists()
            )
            max_score = sum(t.points for t in test_tasks)
            last_attempt = test_attempts.order_by('-timestamp').first()
            grouped_ear[section].append({
                "test": test,
                "score": score,
                "correct_count": correct_count,
                "incorrect_count": incorrect_count,
                "max_score": max_score,
                "last_attempt": last_attempt.timestamp if last_attempt else None,
            })

        sorted_ear_sections = sorted(grouped_ear.keys())
        grouped_ear_training_results = []
        for section in sorted_ear_sections:
            results = grouped_ear[section]
            avg_score = round(sum(r["score"] for r in results) / len(results), 2) if results else 0
            grouped_ear_training_results.append((section, results, avg_score))

        context['grouped_ear_training_results'] = grouped_ear_training_results  # (section, results, avg_score)
        context['grouped_solfeggio_results'] = grouped_solfeggio_list  # (section, results, avg_score)
        context['grouped_literature_results'] = grouped_literature_list  # (section, results, avg_score)
        return context


def test_list(request):
    if not request.user.is_authenticated:
        return render(request, "404.html", status=404)
    # Предполагается, что Lesson, EducationPlan, Subject импортированы из соответствующих модулей
    solfeggio_tests = Test.objects.filter(
        lesson__education_plan__subject__name="Сольфеджио"
    ).select_related('lesson', 'lesson__education_plan', 'lesson__education_plan__subject')

    literature_tests = Test.objects.filter(
        lesson__education_plan__subject__name="Музыкальная литература"
    ).select_related('lesson', 'lesson__education_plan', 'lesson__education_plan__subject')

    context = {
        'solfeggio_tests': solfeggio_tests,
        'literature_tests': literature_tests,
    }
    return render(request, "school/test_list.html", context)

def test_detail(request, slug):
    if not request.user.is_authenticated:
        return render(request, "404.html", status=404)
    test = get_object_or_404(Test, slug=slug)
    return render(request, "school/test_detail.html", {"test": test})

class LiteratureLessonsView(LoginRequiredMixin, ListView):
    model = Lesson
    template_name = 'school/literature_lessons.html'
    context_object_name = 'lessons'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return render(request, "404.html", status=404)
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        user = self.request.user
        # Если пользователь учитель — показываем уроки по мущ. лит. для его школы
        if user.groups.filter(name="Учителя").exists():
            try:
                teacher = Teachers.objects.get(user=user)
            except Teachers.DoesNotExist:
                return Lesson.objects.none()
            school = teacher.school
            # Находим предмет "Сольфеджио"
            try:
                literature_subject = Subject.objects.get(name='Музыкальная литература')
            except Subject.DoesNotExist:
                return Lesson.objects.none()
            # Находим все планы обучения по сольфеджио для школы учителя
            plans = EducationPlan.objects.filter(school=school,subject=literature_subject)
            lessons = Lesson.objects.filter(education_plan__in=plans)
            lessons = lessons.annotate(
                json_id=Cast(F('content_file__id'),output_field=IntegerField())
            ).order_by('json_id')
            return lessons

        # Если пользователь ученик — показываем только его уроки
        try:
            student = Students.objects.get(user=user)
        except Students.DoesNotExist:
            return Lesson.objects.none()
        school = student.school
        grade = student.grade_of_school
        try:
            literature_subject = Subject.objects.get(name='Музыкальная литература')
        except Subject.DoesNotExist:
            return Lesson.objects.none()
        plans = EducationPlan.objects.filter(
            year=grade,
            school=school,
            subject=literature_subject
        )
        lessons = Lesson.objects.filter(education_plan__in=plans)
        lessons = lessons.annotate(
            json_id=Cast(F('content_file__id'),output_field=IntegerField())
        ).order_by('json_id')

        return lessons

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Добавляем информацию о предмете
        try:
            context['subject'] = Subject.objects.get(name='Музыкальная литература')
        except Subject.DoesNotExist:
            pass
        user = self.request.user
        # Если пользователь - ученик, добавляем информацию о его классе
        if user.groups.filter(name="Ученики").exists():
            try:
                student = Students.objects.get(user=user)
                context['student_grade'] = student.grade_of_school
                context['student_school'] = student.school
            except Students.DoesNotExist:
                pass
        # Если пользователь - учитель, добавляем информацию о его школе
        elif user.groups.filter(name="Учителя").exists():
            try:
                teacher = Teachers.objects.get(user=user)
                context['teacher_school'] = teacher.school
            except Teachers.DoesNotExist:
                pass

        return context


class LessonLiteratureDetailView(LoginRequiredMixin, DetailView):
    model = Lesson
    template_name = 'school/lesson_detail_lit.html'
    context_object_name = 'lesson'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return render(request, "404.html", status=404)
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        lesson = self.get_object()
        # Добавляем связанные тесты, если они есть
        context['tests'] = Test.objects.filter(lesson=lesson)
        # Получаем следующий урок
        try:
            context['next_lesson'] = lesson.get_next_lesson()
        except:
            pass

        return context

import io
import openpyxl
from django.http import HttpResponse
from django.views import View
from django import forms
from django.contrib.auth.mixins import LoginRequiredMixin
from django.shortcuts import render
from .models import Lesson, Subject, EducationPlan
from users.models import Students, TestResult
from openpyxl.styles import Border, Side, Alignment
from urllib.parse import quote
from .models import Subject  # убедитесь, что импорт есть
class ExportProgressReportView(LoginRequiredMixin, View):
    template_name = "school/export_progress_report.html"

    def get(self, request):
        user = request.user
        if not user.groups.filter(name="Учителя").exists():
            return render(request, "school/forbidden.html")
        teacher = get_object_or_404(Teachers, user=user)
        school = teacher.school

        grades = (
            Students.objects.filter(school=school)
            .exclude(grade_of_school=None)
            .values_list('grade_of_school', flat=True)
            .distinct()
            .order_by('grade_of_school')
        )

        form = ProgressReportExportForm(school=school, teacher=teacher)
        return render(request, self.template_name, {"form": form, "grades": grades})

    def post(self, request):
        user = request.user
        if not user.groups.filter(name="Учителя").exists():
            return render(request, "school/forbidden.html")
        teacher = get_object_or_404(Teachers, user=user)
        school = teacher.school

        grades = (
            Students.objects.filter(school=school)
            .exclude(grade_of_school=None)
            .values_list('grade_of_school', flat=True)
            .distinct()
            .order_by('grade_of_school')
        )

        form = ProgressReportExportForm(request.POST, school=school, teacher=teacher)
        selected_grade = request.POST.get("grade_of_school")
        if form.is_valid():
            section = form.cleaned_data.get('section')
            student = form.cleaned_data.get('student')
            group = form.cleaned_data.get('group')
            subject = form.cleaned_data.get('subject')

            # Фильтрация уроков
            lesson_qs = Lesson.objects.filter(education_plan__school=school)
            if subject:
                lesson_qs = lesson_qs.filter(education_plan__subject=subject)
            if section:
                lesson_qs = lesson_qs.filter(section=section)
            if selected_grade and str(selected_grade).isdigit():
                lesson_qs = lesson_qs.filter(education_plan__year=int(selected_grade))
            lessons = lesson_qs

            # Фильтрация учеников с сортировкой по фамилии и имени
            if group:
                students_qs = Students.objects.filter(
                    pk__in=GroupMembership.objects.filter(group=group).values_list('student_id', flat=True)
                )
            elif student:
                students_qs = Students.objects.filter(pk=student.pk)
            else:
                students_qs = Students.objects.filter(school=school)
                if selected_grade and str(selected_grade).isdigit():
                    students_qs = students_qs.filter(grade_of_school=int(selected_grade))
            students_qs = students_qs.order_by('surname', 'name')

            # Получаем результаты тестов
            test_results = TestResult.objects.filter(
                student__in=students_qs,
                test__lesson__in=lessons
            ).select_related('student', 'test', 'test__lesson')

            # Получаем результаты контрольных по слуховому анализу только если выбран предмет "Сольфеджио"
            ear_rows = []
            solfedgio_subject = Subject.objects.filter(name__iexact="Сольфеджио").first()
            if subject and solfedgio_subject and subject.pk == solfedgio_subject.pk:
                ear_tests = EarTrainingTest.objects.all()
                if section:
                    ear_tests = ear_tests.filter(section=section)
                if selected_grade and str(selected_grade).isdigit():
                    ear_tests = ear_tests.filter(lesson__education_plan__year=int(selected_grade))
                # Сортируем студентов по фамилии и имени
                for st in students_qs:
                    for etest in ear_tests:
                        tasks = etest.tasks.all()
                        attempts = EarTrainingAttempt.objects.filter(user=st.user, task__in=tasks)
                        score = sum(
                            t.points for t in tasks
                            if attempts.filter(task=t, is_correct=True).exists()
                        )
                        max_score = sum(t.points for t in tasks)
                        correct_count = attempts.filter(is_correct=True).count()
                        incorrect_count = attempts.filter(is_correct=False).count()
                        last_attempt = attempts.order_by('-timestamp').first()
                        if attempts.exists():
                            ear_rows.append([
                                st.surname,
                                st.name,
                                etest.section,
                                "Сольфеджио (контрольная)",
                                etest.lesson.topic if etest.lesson else "",
                                etest.title,
                                score,
                                max_score,
                                "-",  # Попытка (можно реализовать по-другому)
                                last_attempt.timestamp.strftime("%d.%m.%Y %H:%M") if last_attempt else "",
                            ])
                # Сортировка ear_rows по фамилии и имени ученика
                ear_rows.sort(key=lambda row: (row[0], row[1]))

            # Генерация Excel
            output = io.BytesIO()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Отчет по успеваемости"

            # Заголовки
            headers = [
                "Фамилия", "Имя", "Секция", "Предмет", "Тема урока", "Тема теста",
                "Баллы", "Максимум", "Попытка", "Дата"
            ]
            ws.append(headers)

            # Основные тесты (сортировка по фамилии и имени)
            for res in test_results.order_by('student__surname', 'student__name', 'test__lesson__section', 'test__topic'):
                ws.append([
                    res.student.surname,
                    res.student.name,
                    res.test.lesson.section,
                    res.test.lesson.education_plan.subject.name,
                    res.test.lesson.topic,
                    res.test.topic,
                    res.score,
                    res.max_score,
                    res.attempt_number,
                    res.timestamp.strftime("%d.%m.%Y %H:%M"),
                ])

            # Контрольные по сольфеджио (только если выбран предмет "Сольфеджио")
            for row in ear_rows:
                ws.append(row)

            # --- Стилизация: границы, перенос текста, автоширина ---
            thin = Side(border_style="thin", color="000000")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            wrap_alignment = Alignment(wrap_text=True, vertical="center")

            # Применяем стили ко всем ячейкам
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = border
                    cell.alignment = wrap_alignment

            # Автоматическая ширина столбцов
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        cell_value = str(cell.value) if cell.value is not None else ""
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = max(12, min(max_length + 2, 50))

            wb.save(output)
            output.seek(0)
            if group:
                filename = f"Группа_{group.name}_{section or ''}_{subject or ''}_{selected_grade or ''}.xlsx"
            elif student:
                filename = f"{student.surname} {student.name} {section or ''} {subject or ''} {selected_grade or ''}.xlsx"
            else:
                filename = f"{school.name} {section or ''} {subject or ''} {selected_grade or ''} отчет.xlsx"
            response = HttpResponse(
                output,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            filename_ascii = filename.encode('ascii', 'ignore').decode() or "report.xlsx"
            filename_utf8 = quote(filename)

            response['Content-Disposition'] = (
                f'attachment; filename="{filename_ascii}"; filename*=UTF-8\'\'{filename_utf8}'
            )
            return response

        return render(request, self.template_name, {"form": form, "grades": grades})



from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse, Http404
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from .models import Lesson, EarTrainingTest, EarTrainingTask, EarTrainingAttempt
import json

def is_student(user):
    return user.groups.filter(name='Ученики').exists()

@login_required
@user_passes_test(is_student)
def ear_training_tests(request):
    # Получаем текущего ученика
    try:
        student = Students.objects.get(user=request.user)
    except Students.DoesNotExist:
        return render(request, 'school/ear_training_tests.html', {'sections': {}})

    # Получаем школу и год обучения ученика
    school = student.school
    year = student.grade_of_school

    # Фильтруем тесты по школе и году обучения через Lesson -> EducationPlan
    tests = (
        EarTrainingTest.objects
        .select_related('lesson', 'lesson__education_plan')
        .filter(
            lesson__education_plan__school=school,
            lesson__education_plan__year=year
        )
        .order_by('section', 'lesson__topic')
    )

    # Группируем тесты по section
    sections = {}
    for test in tests:
        sections.setdefault(test.section, []).append(test)

    return render(request, 'school/ear_training_tests.html', {'sections': sections})


@login_required
@user_passes_test(is_student)
def ear_training_test_detail(request, test_id):
    test = get_object_or_404(EarTrainingTest, id=test_id)
    tasks = test.tasks.all().order_by('id')
    # Проверяем, не завершил ли пользователь уже все задания
    attempts = EarTrainingAttempt.objects.filter(user=request.user, task__in=tasks)
    completed_task_ids = set(a.task_id for a in attempts if a.is_correct)
    # Найти первый незавершённый task
    for task in tasks:
        if task.id not in completed_task_ids:
            return redirect('school:ear_training_task', test_id=test.id, task_id=task.id)
    # Все задания завершены — показать результат
    return redirect('school:ear_training_test_result', test_id=test.id)

@login_required
@user_passes_test(is_student)
def ear_training_task(request, test_id, task_id):
    test = get_object_or_404(EarTrainingTest, id=test_id)
    task = get_object_or_404(EarTrainingTask, id=task_id, test=test)
    # Проверяем, не исчерпаны ли попытки
    attempts = EarTrainingAttempt.objects.filter(user=request.user, task=task)
    max_attempts = 2 if task.task_type == 'performance' else 1
    if attempts.filter(is_correct=True).exists() or attempts.count() >= max_attempts:
        # Перейти к следующему заданию
        tasks = list(test.tasks.all().order_by('id'))
        idx = tasks.index(task)
        if idx + 1 < len(tasks):
            return redirect('school:ear_training_task', test_id=test.id, task_id=tasks[idx+1].id)
        else:
            return redirect('school:ear_training_test_result', test_id=test.id)
    return render(request, 'school/ear_training_task.html', {'task': task, 'test': test, 'attempts_left': max_attempts - attempts.count()})

@csrf_exempt
@login_required
@user_passes_test(is_student)
def submit_attempt(request):
    if request.method != 'POST':
        raise Http404
    data = json.loads(request.body)
    task_id = data.get('task_id')
    task = get_object_or_404(EarTrainingTask, id=task_id)
    user = request.user
    attempts = EarTrainingAttempt.objects.filter(user=user, task=task)
    max_attempts = 2 if task.task_type == 'performance' else 1
    if attempts.filter(is_correct=True).exists() or attempts.count() >= max_attempts:
        return JsonResponse({'result': 'max_attempts', 'message': 'Попытки закончились.'})
    if task.task_type == 'recognition':
        selected_option = int(data.get('selected_option'))
        is_correct = (selected_option == task.correct_option)
        EarTrainingAttempt.objects.create(
            user=user,
            task=task,
            attempt_number=attempts.count() + 1,
            selected_option=selected_option,
            is_correct=is_correct
        )
        return JsonResponse({
            'result': 'ok',
            'is_correct': is_correct,
            'attempts_left': max_attempts - attempts.count() - (1 if is_correct else 0)
        })
    elif task.task_type == 'performance':
        played_notes = data.get('played_notes', [])
        ref_notes = [n for n in task.reference_notes]
        is_correct = (played_notes == ref_notes)
        EarTrainingAttempt.objects.create(
            user=user,
            task=task,
            attempt_number=attempts.count() + 1,
            played_notes=played_notes,
            is_correct=is_correct
        )
        return JsonResponse({
            'result': 'ok',
            'is_correct': is_correct,
            'attempts_left': max_attempts - attempts.count() - (1 if is_correct else 0)
        })

@login_required
@user_passes_test(is_student)
def ear_training_test_result(request, test_id):
    test = get_object_or_404(EarTrainingTest, id=test_id)
    tasks = test.tasks.all()
    # Получаем все попытки пользователя по заданиям этой контрольной
    attempts = EarTrainingAttempt.objects.filter(user=request.user, task__in=tasks)
    correct = attempts.filter(is_correct=True).count()
    total = tasks.count()
    # Готовим словарь: task_id -> [список попыток]
    attempts_by_task = {}
    for attempt in attempts:
        attempts_by_task.setdefault(attempt.task_id, []).append(attempt)
    # Готовим список пар (task, attempts) для шаблона
    tasks_and_attempts = []
    for task in tasks:
        task_attempts = attempts_by_task.get(task.id, [])
        tasks_and_attempts.append((task, task_attempts))
    return render(request, 'school/ear_training_test_result.html', {
        'test': test,
        'correct': correct,
        'total': total,
        'attempts_by_task': attempts_by_task,
        'tasks_and_attempts': tasks_and_attempts,
    })

@login_required
@user_passes_test(is_student)
def ear_training_results_section(request, section_name):
    tests = EarTrainingTest.objects.filter(section=section_name)
    results = []
    for test in tests:
        tasks = test.tasks.all()
        attempts = EarTrainingAttempt.objects.filter(user=request.user, task__in=tasks)
        correct = attempts.filter(is_correct=True).count()
        total = tasks.count()
        results.append({
            'test': test,
            'lesson': test.lesson,
            'correct': correct,
            'total': total,
            'status': 'Сдано' if correct == total else 'Не сдано',
        })
    return render(request, 'school/ear_training_results_section.html', {
        'section_name': section_name,
        'results': results,
    })


from collections import defaultdict
from django.db.models import Sum, Count, Max

